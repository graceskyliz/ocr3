from openpyxl import load_workbook
from io import BytesIO

def parse_invoice_xlsx(raw: bytes):
    wb = load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    text = {}
    for r in ws.iter_rows(max_row=min(ws.max_row, 200), max_col=min(ws.max_column, 20)):
        for c in r:
            v = str(c.value).strip() if c.value is not None else ""
            if v:
                text[(c.row, c.column)] = v.lower()

    def find_right(lbl):
        for (rr, cc), val in text.items():
            if lbl in val:
                right = ws.cell(rr, cc+1).value
                return str(right).strip() if right is not None else ""
        return ""

    fields = {}
    for key in ["ruc","proveedor","fecha","subtotal","igv","total","moneda","serie","numero"]:
        fields[key] = find_right(key)
    return {"fields": fields, "items": [], "confidence": None}
